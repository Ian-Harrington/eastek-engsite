import os
import types
import math
import datetime
from io import BytesIO

from django.conf import settings

#from reportlab.pdfgen import canvas
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook, worksheet
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image

RECORDS_PER_SHEET = 32 



def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
	""" Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
	This is monkeypatched to remove cell deletion bug
	https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
	"""
	if not range_string and not all((start_row, start_column, end_row, end_column)):
		msg = "You have to provide a value either for 'coordinate' or for\
		'start_row', 'start_column', 'end_row' *and* 'end_column'"
		raise ValueError(msg)
	elif not range_string:
		range_string = '%s%s:%s%s' % (get_column_letter(start_column),
											start_row,
										get_column_letter(end_column),
										end_row)
	elif ":" not in range_string:
		if COORD_RE.match(range_string):
			return  # Single cell, do nothing
		raise ValueError("Range must be a cell range (e.g. A1:E1)")
	else:
		range_string = range_string.replace('$', '')

	if range_string not in self._merged_cells:
		self._merged_cells.append(range_string)


# Apply monkey patch
merge_cells = types.MethodType(merge_cells, None, worksheet.Worksheet)
worksheet.Worksheet.merge_cells = merge_cells


"""
def simpleExample():
	# stolen from django overview
	buffer = BytesIO()
	# Create the PDF object, using the BytesIO object as its "file."
	p = canvas.Canvas(buffer)
	# Draw things on the PDF. Here's where the PDF generation happens.
	# See the ReportLab documentation for the full list of functionality.
	p.drawString(100, 100, "Hello world.")
	# Close the PDF object cleanly.
	p.showPage()
	p.save()
	# Get the value of the BytesIO buffer and write it to the response.
	pdf = buffer.getvalue()
	buffer.close()
	return pdf
"""

# accepts iterable of overtime objects, returns pdf
def genOTpdf(overtime): 
	otdate = overtime[0].date
	prepOT(overtime)
	# Parse the overtime info into a list of list of required OT info
	# pass LoLs & date to pdfgen
	#	create document 
	#	create template
	#	loop over LoLs
	#		writes info to flowables
	#	break every 34(?) entries to different pages
	#	close & save

# accepts an overtime object, returns list of strings to be written	
def prepOT(ot):
	OTstr = []
	for o in ot:
		OTstr.append[o.emp.emp_id]
		OTstr.append[o.emp.name]
		OTstr.append[str(o.time) + '~' + str(o.time) + datetime.time(o.request_hours)]
		OTstr.append[str(o.time) + '~' + str(o.time) + datetime.time(o.request_hours)]
		OTstr.append[str(o.request_hours) + 'H']
		OTstr.append[str(o.reason)]
	return OTstr


# accepts iterable of overtime objects, returns list of list of strings (same as in access output)
def OTlines(overtime):
	ot = [['Employee ID', 'Name', 'Hours', 'Customer', 'Project', 'Start Time', 'Date', 'Reason'],]
	for o in overtime:
		ot.append([o.emp.emp_id, o.emp.name, o.request_hours, o.project.customer, o.project.project,
			o.time, o.date, o.reason])
	return ot


# accepts iterable of overtime objects, returns virtual workbook
def generate_overtime_report(overtime):
	#wb = Workbook(encoding='utf-8') # open a workbook which serves as a template
	wb = load_workbook(os.path.join(settings.BASE_DIR, 'overtime/static/OT_Request_Template.xlsx')) # don't exactly know how to do this
	# set up the date for the template
	date = overtime[0].date
	ws = wb['Template']
	ws['H4'].value = ws['H4'].value + str(date.year) + ws['J4'].value + str(date.month) + ws['K4'].value + str(date.day) + ws['L4'].value
	ws['J4'] = ws['K4'] = ws['L4'] = None
	img = Image(os.path.join(settings.BASE_DIR, 'EngSite/static/img/eastek_logo.png'))
	for sheet in range(1, math.ceil(len(overtime)/RECORDS_PER_SHEET)+1):
		ws = wb.copy_worksheet(wb['Template']) # copy the template worksheet
		for i in range(1, min(RECORDS_PER_SHEET, len(overtime))+1):
			ot = overtime[i*sheet-1]
			img.anchor(ws['A1'])
			ws.add_image(img)
			ws.cell(row=i+5, column=1, value=i*sheet) # index (value might be just i)
			ws.cell(row=i+5, column=2, value=ot.emp.name)
			ws.cell(row=i+5, column=3, value=ot.emp.emp_id)
			duration = ot.time.strftime('%H:%M') + '~' + (datetime.datetime(100,1,1,ot.time.hour,ot.time.minute,0) + datetime.timedelta(hours=float(ot.request_hours))).strftime('%H:%M')
			ws.cell(row=i+5, column=5, value=duration)
			ws.cell(row=i+5, column=6, value=duration)

			ws.cell(row=i+5, column=7, value=str(max(int(ot.request_hours), ot.request_hours)) + 'H')
			ws.cell(row=i+5, column=8, value=ot.project.name + ' - ' + ot.reason)
	wb.remove_sheet(wb['Template'])
	return save_virtual_workbook(wb)

# accepts a worksheet object and a date, returns the worksheet
def generate_xlsx_template(ws, date):

	return ws
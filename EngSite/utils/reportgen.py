import os
import types
import math
import datetime
from io import BytesIO

from django.conf import settings

from openpyxl.utils import range_boundaries
from openpyxl import load_workbook, worksheet
from openpyxl.styles import Border, Side
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image

from .utils import choice_lookup
from projects.models import ChecklistItem

RECORDS_PER_SHEET = 29 


def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
	""" Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
	This is monkeypatched to remove cell deletion bug
	https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
	"""
	if not range_string and not all((start_row, start_column, end_row, end_column)):
		msg = "You have to provide a value either for 'coordinate' or for\
		'start_row', 'start_column', 'end_row' *and* 'end_column'"
		raise ValueError(msg)
	elif not range_string:
		range_string = '%s%s:%s%s' % (get_column_letter(start_column),
										start_row,
										get_column_letter(end_column),
										end_row)
	elif ":" not in range_string:
		if COORD_RE.match(range_string):
			return  # Single cell, do nothing
		raise ValueError("Range must be a cell range (e.g. A1:E1)")
	else:
		range_string = range_string.replace('$', '')

	if range_string not in self._merged_cells:
		self._merged_cells.append(range_string)


# Apply monkey patch
# My method (because *all* the code here is run *after* importing the class but before instancing)
#     if openpyxl is used elsewhere on workbooks with merged cells the patch may need to be run again (IDK)
worksheet.Worksheet.merge_cells = merge_cells


# accepts iterable of overtime objects, returns virtual workbook
def generate_overtime_report(overtime):
	wb = load_workbook(os.path.join(settings.BASE_DIR, 'overtime/static/OT_Request_Template.xlsx'))
	# set up the date for the template
	date = overtime[0].date
	ws = wb['Template']
	ws['H4'].value = ws['H4'].value + str(date.year) + ws['J4'].value + str(date.month) + ws['K4'].value + str(date.day) + ws['L4'].value
	ws['J4'] = ws['K4'] = ws['L4'] = None
	img = Image(os.path.join(settings.BASE_DIR, 'EngSite/static/img/eastek_logo.png'))
	for sheet in range(1, math.ceil(len(overtime)/RECORDS_PER_SHEET)+1):
		ws = wb.copy_worksheet(wb['Template']) # copy the template worksheet
		for i in range(1, min(RECORDS_PER_SHEET, len(overtime))+1):
			ot = overtime[i*sheet-1]
			img.anchor(ws['A1'])
			ws.add_image(img)
			ws.cell(row=i+5, column=1, value=i*sheet) # index (value might be just i)
			ws.cell(row=i+5, column=2, value=ot.emp.name)
			ws.cell(row=i+5, column=3, value=ot.emp.emp_id)
			duration = ot.time.strftime('%H:%M') + '~' + (datetime.datetime(100,1,1,ot.time.hour,ot.time.minute,0) + datetime.timedelta(hours=float(ot.request_hours))).strftime('%H:%M')
			ws.cell(row=i+5, column=5, value=duration)
			ws.cell(row=i+5, column=6, value=duration)
			ws.cell(row=i+5, column=7, value=str(max(int(ot.request_hours), ot.request_hours)) + 'H')
			ws.cell(row=i+5, column=8, value=ot.project.name + ' - ' + ot.reason)
	wb.remove_sheet(wb['Template'])
	return save_virtual_workbook(wb)

def generate_gate_checklist(ms):
	wb = load_workbook(os.path.join(settings.BASE_DIR, 'projects/static/Checklist_Template.xlsx'))
	ws = wb['Template']
	ws.title = ms.description
	# Header (any information there is)
	ws['A1'].value = ms.description + ' Checklist'
	ws['C3'].value = ms.project.customer.name
	ws['C4'].value = ms.project.name
	ws['H2'].value = str(ms.completion_date)
	ws['H3'].value = ms.project.eastek_pn
	ws['H4'].value = ms.project.cust_pn
	#ws['H5'].value = ms.project.description # if a description is added to project model
	ws['H6'].value = ms.project.engineer.all()[0].english_name
	# Checklist
	t = ms.checklist.count() - 1
	for i in range(ms.checklist.count()): #checklist is a queryset
		ws.cell(row=i+8, column=1, value=i+1)
		ws.cell(row=i+8, column=2, value=ms.checklist.all()[i].name)
		ws.cell(row=i+8, column=4, value=choice_lookup(opt_list=ChecklistItem.RESPONSIBLE, item=ms.checklist.all()[i].responsible))
		if ms.checklist.all()[i].completed == True:
			ws.cell(row=i+8, column=5, value='X')
		elif ms.checklist.all()[i].completed == False:
			ws.cell(row=i+8, column=6, value='X')
		elif ms.checklist.all()[i].completed == None:
			ws.cell(row=i+8, column=7, value='X')
		ws.cell(row=i+8, column=8, value=ms.checklist.all()[i].remarks)
	# remove formatting from all cells below
	# add the thick end line
	no_border = Border(left=Side(border_style=None),
						right=Side(border_style=None),
						top=Side(border_style=None),
						bottom=Side(border_style=None))
	strong_top = Border(top=Side(border_style='medium'))
	for row in ws.iter_rows(min_row=i+9, max_col=8, max_row=64):
		for c in row:
			c.border = no_border
	for c in ws[i+9]:
		c.border = strong_top
	return save_virtual_workbook(wb)
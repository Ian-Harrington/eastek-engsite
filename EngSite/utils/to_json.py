#!usr/bin/env python

from openpyxl import load_workbook
from datetime import date
import csv
import sys
import re

def CSV_to_JSON(file_path, obj_name):
# open csv
	with open(file_path) as file:
		rdr = csv.reader(file)
		header = rdr.__next__()
		objects = []
		for r in rdr:
			objects.append(lstr_to_JSON(obj_name=obj_name, header=header, obj=r))
	# write to file
	fName = file_path.split('.')[0] + '_to_JSON.json'
	with open(fName, 'w') as f:
		f.write('[')
		for o in objects:
			f.write(o + ",")
		f.write(']')

#converts each line into a JSON object
def lstr_to_JSON(obj_name, header, obj):
	# handles the model, primary key, and fields opener
	JSON_str = '{"model": "' + obj_name + '", "pk": ' + parse(obj[0]) + ', "fields": {'
	for i in range(1,len(obj)):
		JSON_str = JSON_str + '"' + header[i] + '": ' + parse(obj[i])
		if i != len(obj)-1:
			JSON_str = JSON_str + ','
	JSON_str = JSON_str + '}}'
	return JSON_str


# parse True/False, Numbers, Dates, Blank, Strings
def parse(data):
	tmp = data
	if data.lower() == 'true' or data.lower() == '=true()':
		return 'true'
	elif data.lower() == 'false' or data.lower() == '=false()':
		return 'false'
	elif re.match(r'^[1-9][0-9]*(\.)?[0-9]*$', data): # doesn't catch 0.5 or .5 (maybe not a big deal)
		#isinstance(data, type(1)) or isinstance(data, type(1.0)): # all ints or floats (xlsx breaks this)
		return tmp
	elif re.match(r'^[12][0-9]{3}-[01][0-9]-[0-3][0-9]', data):
		return '"' + re.match(r'^[12][0-9]{3}-[01][0-9]-[0-3][0-9]', data).group() + '"'
	elif data == 'None' or data == '': # is the second one needed or good?
		return 'null'
	else:
		return '"' + tmp + '"'

# sensitive to files with cells which contain null data (CRTL+> as though they have a value)
def XLSX_to_JSON(file_path, obj_name):
	# opens the workbook
	wb = load_workbook(filename=file_path, read_only=True)
	# gets the active worksheet if there is no worksheet named the as the model
	for sht in wb:
		if sht.title == obj_name.split('.')[1]:
			ws = sht
			break
	else:
		ws = wb.active
	header = []
	# grabs the first line (header)
	for row in ws.iter_rows(min_row=1, max_row=1):
		for c in row:
			header.append(str(c.value))
	# get the rest of the rows
	# ignore the first row
	objs = []
	for row in tuple(ws.rows)[1:]: # use slice to avoid headers
		rVal = []
		# get string values for each cell in the row
		for c in row:
			rVal.append(str(c.value))
		objs.append(lstr_to_JSON(obj_name=obj_name, header=header, obj=rVal))
	# opens a new file to write to & writes lines
	fName = obj_name.split('.')[1] + date.today().strftime('%Y-%m-%d') + '.json'
	with open(fName, 'w') as f:
		f.write('[')
		f.write(objs.pop(0))
		for o in objs:
			f.write(', ' + o)
		f.write(']')


if __name__ == '__main__':
	if sys.argv[1].split('.')[1].lower() == 'csv':
		CSV_to_JSON(sys.argv[1], sys.argv[2])
	elif sys.argv[1].split('.')[1].lower() == 'xlsx':
		XLSX_to_JSON(sys.argv[1], sys.argv[2])
	else:
		print('Unrecognized file type')